#!/usr/bin/env python3
"""Screen source-data tables for possible unit-of-analysis mismatches."""

from __future__ import annotations

import argparse
import csv
import json
import math
from collections import defaultdict
from pathlib import Path
from typing import Any


CSV_EXTS = {".csv", ".tsv"}
XLSX_EXTS = {".xlsx"}
TABLE_EXTS = CSV_EXTS | XLSX_EXTS
BIOLOGICAL_ID_COLUMNS = (
    "animal",
    "animal_id",
    "mouse",
    "mouse_id",
    "rat",
    "rat_id",
    "subject",
    "subject_id",
    "patient",
    "patient_id",
    "participant",
    "participant_id",
    "donor",
    "donor_id",
)
TECHNICAL_ID_COLUMNS = (
    "field_id",
    "field_num",
    "field_number",
    "section_id",
    "section_num",
    "section_number",
    "well_id",
    "well_num",
    "well_number",
    "technical_replicate",
    "technical_replicate_id",
    "technical_replicate_num",
    "technical_replicate_number",
    "replicate_id",
    "replicate_num",
    "replicate_number",
    "cell_id",
    "lesion_id",
    "image_id",
    "image_num",
    "image_number",
    "visit",
    "visit_id",
    "timepoint",
    "timepoint_id",
)


def normalize_header(header: str) -> str:
    return header.strip().lower().replace(" ", "_").replace("-", "_")


def cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def xlsx_header_score(values: list[Any]) -> float:
    populated = [cell_to_text(value).strip() for value in values if cell_to_text(value).strip()]
    if len(populated) < 2:
        return float("-inf")
    normalized = [normalize_header(value) for value in populated]
    hints = (
        set(BIOLOGICAL_ID_COLUMNS)
        | set(TECHNICAL_ID_COLUMNS)
        | {"group", "condition", "treatment", "arm", "value", "reported_n_basis"}
    )
    recognized = sum(value in hints for value in normalized)
    numeric = 0
    for value in populated:
        try:
            float(value)
            numeric += 1
        except ValueError:
            pass
    return (recognized * 8.0) + ((len(populated) - numeric) * 1.5) - (numeric * 3.0)


def select_xlsx_header_index(matrix: list[list[Any]], search_rows: int = 20) -> int | None:
    nonempty = [idx for idx, values in enumerate(matrix[:search_rows]) if any(cell_to_text(value).strip() for value in values)]
    if not nonempty:
        return None
    scored = [(xlsx_header_score(matrix[idx]), idx) for idx in nonempty]
    best_score, best_idx = max(scored, key=lambda item: (item[0], -item[1]))
    return best_idx if math.isfinite(best_score) else nonempty[0]


def read_delimited_table(path: Path) -> list[dict[str, str]]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    with path.open(newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh, delimiter=delimiter)
        return [{normalize_header(k): v for k, v in row.items() if k is not None} for row in reader]


def read_xlsx_tables(path: Path) -> list[tuple[Path, list[dict[str, str]]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # noqa: BLE001 - xlsx support is an explicit dependency.
        raise RuntimeError("XLSX pseudoreplication screening requires openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    tables: list[tuple[Path, list[dict[str, str]]]] = []
    for sheet in workbook.worksheets:
        matrix = [list(values) for values in sheet.iter_rows(values_only=True)]
        header_idx = select_xlsx_header_index(matrix)
        if header_idx is None:
            continue
        headers = [
            normalize_header(cell_to_text(value)) if cell_to_text(value).strip() else f"column_{idx + 1}"
            for idx, value in enumerate(matrix[header_idx])
        ]
        rows: list[dict[str, str]] = []
        for values in matrix[header_idx + 1:]:
            if not any(cell_to_text(value).strip() for value in values):
                continue
            rows.append({
                header: cell_to_text(value)
                for header, value in zip(headers, values)
                if header
            })
        if rows:
            tables.append((Path(f"{path.name}#{sheet.title}"), rows))
    workbook.close()
    return tables


def read_tables(path: Path) -> list[tuple[Path, list[dict[str, str]]]]:
    if path.suffix.lower() in XLSX_EXTS:
        return read_xlsx_tables(path)
    return [(path, read_delimited_table(path))]


def collect_files(path: Path) -> list[Path]:
    if path.is_file():
        return [path] if path.suffix.lower() in TABLE_EXTS else []
    return [
        p for p in sorted(path.rglob("*"))
        if not p.is_symlink() and p.is_file() and p.suffix.lower() in TABLE_EXTS
    ]


def present_column(rows: list[dict[str, str]], candidates: tuple[str, ...]) -> str | None:
    if not rows:
        return None
    columns = set(rows[0])
    for candidate in candidates:
        if candidate in columns:
            return candidate
    return None


def group_column(rows: list[dict[str, str]]) -> str | None:
    for candidate in ("group", "condition", "treatment", "arm"):
        if rows and candidate in rows[0]:
            return candidate
    return None


def screen_table(path: Path, rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    if not rows:
        return []
    biological = present_column(rows, BIOLOGICAL_ID_COLUMNS)
    technical = present_column(rows, TECHNICAL_ID_COLUMNS)
    if not biological or not technical:
        return []

    group_col = group_column(rows)
    buckets: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        key = row.get(group_col, "all") if group_col else "all"
        buckets[key].append(row)

    candidates: list[dict[str, Any]] = []
    for group, group_rows in buckets.items():
        biological_units = {row.get(biological, "") for row in group_rows if row.get(biological, "")}
        technical_units = {(row.get(biological, ""), row.get(technical, "")) for row in group_rows if row.get(technical, "")}
        if len(biological_units) >= 1 and len(technical_units) > len(biological_units):
            reported_n_basis_values = {row.get("reported_n_basis", "").lower() for row in group_rows}
            reported_technical = any(value in {"field", "fields", "well", "wells", "technical", "cell", "cells"} for value in reported_n_basis_values)
            candidate_id = f"STAT-PSEUDO-{len(candidates) + 1:04d}"
            candidates.append({
                "candidate_id": candidate_id,
                "detector": "stats.pseudoreplication_screen",
                "candidate_type": "pseudoreplication_candidate",
                "locations": [f"{path.name}:group={group}"],
                "evidence": {
                    "file": str(path),
                    "group": group,
                    "biological_id_column": biological,
                    "technical_id_column": technical,
                    "biological_unit_count": len(biological_units),
                    "technical_unit_count": len(technical_units),
                    "row_count": len(group_rows),
                    "reported_n_basis_values": sorted(v for v in reported_n_basis_values if v),
                    "reported_n_appears_technical": reported_technical,
                },
                "evidence_strength": "weak_signal",
                "risk_suggestion": "R2_possible" if reported_technical else "R1_possible",
                "risk_cap_tags": ["pseudoreplication_candidate", "weak_statistical_signal"],
                "benign_explanations": [
                    "analysis may use a nested or mixed-effects model",
                    "technical replicates may have been averaged before inferential testing",
                    "reported n may be descriptive rather than inferential",
                ],
                "required_materials": [
                    "analysis code",
                    "statistical model specification",
                    "raw measurements by biological unit",
                    "reported n definition from methods or legend",
                ],
                "recommended_action": "Verify whether inferential n counts biological units; reanalyse at the biological-unit level or justify a nested model.",
                "requires_contextual_calibration": True,
            })
    return candidates


def scan(root: Path) -> dict[str, Any]:
    files = collect_files(root)
    candidates: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    for file_path in files:
        try:
            for table_path, rows in read_tables(file_path):
                candidates.extend(screen_table(table_path, rows))
        except Exception as exc:  # noqa: BLE001
            errors.append({"path": str(file_path), "error": str(exc)})
    for idx, candidate in enumerate(candidates, start=1):
        candidate["candidate_id"] = f"STAT-PSEUDO-{idx:04d}"
    return {
        "detector_name": "stats.pseudoreplication_screen",
        "detector_version": "0.3.1",
        "input": {"root": str(root)},
        "candidates": candidates,
        "errors": errors,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("path", type=Path)
    parser.add_argument("--output", type=Path, default=Path("pseudoreplication_candidates.json"))
    args = parser.parse_args()

    root = args.path.expanduser().resolve()
    result = scan(root)
    args.output.write_text(json.dumps(result, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({
        "output": str(args.output),
        "candidates": len(result["candidates"]),
        "errors": len(result["errors"]),
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
