#!/usr/bin/env python3
"""Index XLSX workbook and sheet structure for material preparation.

This is an intake helper, not a statistical detector. It records workbook
sheet names, header-like rows, row/column dimensions, formula/merged-cell
counts, and figure/table-like labels so teams can prepare claim manifests and
decide which sheets need CSV exports. It does not validate calculations or
prove that plotted values match source data.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
import re
from typing import Any


XLSX_EXTS = {".xlsx"}
MAX_ROWS_SCANNED = 1000
MAX_COLUMNS_SCANNED = 64
MAX_HEADERS = 24
FIGURE_TABLE_LABEL_RE = re.compile(
    r"\b((?:supplementary\s+|extended\s+data\s+)?(?:fig(?:ure)?\.?|table)\s*[A-Za-z0-9][A-Za-z0-9.\-]*)\b",
    re.I,
)


def collect_xlsx_files(root: Path) -> list[Path]:
    return sorted(path for path in root.rglob("*") if path.is_file() and path.suffix.lower() in XLSX_EXTS)


def cell_to_display_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if len(text) > 160:
        return text[:157] + "..."
    return text


def suggested_xlsx_label(path: str, sheet_name: str) -> str:
    sheet = str(sheet_name).strip()
    match = FIGURE_TABLE_LABEL_RE.search(sheet)
    if match:
        return " ".join(match.group(1).replace("Fig.", "Figure").split())
    stem = Path(path).stem.replace("_", " ").replace("-", " ")
    match = FIGURE_TABLE_LABEL_RE.search(stem)
    if match:
        return " ".join(match.group(1).replace("Fig.", "Figure").split())
    return sheet


def label_looks_claim_like(label: str) -> bool:
    return bool(FIGURE_TABLE_LABEL_RE.search(label))


def table_payload(table: Any) -> dict[str, str]:
    return {
        "name": str(getattr(table, "name", "") or getattr(table, "displayName", "") or ""),
        "display_name": str(getattr(table, "displayName", "") or ""),
        "ref": str(getattr(table, "ref", "") or ""),
    }


def scan_sheet(path: str, worksheet: Any) -> dict[str, Any]:
    header_row = 0
    headers: list[str] = []
    data_rows_scanned = 0
    non_empty_rows_scanned = 0
    formula_cell_count = 0
    non_empty_cell_count = 0
    row_scan_capped = False
    column_scan_capped = int(worksheet.max_column or 0) > MAX_COLUMNS_SCANNED

    max_row = int(worksheet.max_row or 0)
    max_column = int(worksheet.max_column or 0)
    columns_to_scan = min(max_column, MAX_COLUMNS_SCANNED)
    for row_index, row in enumerate(
        worksheet.iter_rows(
            min_row=1,
            max_row=min(max_row, MAX_ROWS_SCANNED),
            min_col=1,
            max_col=columns_to_scan,
            values_only=False,
        ),
        start=1,
    ):
        cells = [cell_to_display_text(cell.value) for cell in row]
        if any(cell.strip() for cell in cells):
            non_empty_rows_scanned += 1
        for cell in row:
            text = cell_to_display_text(cell.value)
            if text:
                non_empty_cell_count += 1
            if text.startswith("="):
                formula_cell_count += 1
        if not any(cell.strip() for cell in cells):
            continue
        if not headers:
            header_row = row_index
            headers = [
                cell.strip() if cell.strip() else f"column_{idx + 1}"
                for idx, cell in enumerate(cells[:MAX_HEADERS])
            ]
            continue
        data_rows_scanned += 1
    if max_row > MAX_ROWS_SCANNED:
        row_scan_capped = True

    tables = [table_payload(table) for table in getattr(worksheet, "tables", {}).values()]
    merged_ranges = [str(item) for item in list(getattr(worksheet.merged_cells, "ranges", []) or [])[:50]]
    sheet_name = str(worksheet.title)
    label = suggested_xlsx_label(path, sheet_name)
    return {
        "source_xlsx": path,
        "sheet_name": sheet_name,
        "suggested_label": label,
        "looks_figure_or_table_like": label_looks_claim_like(label),
        "sheet_state": str(getattr(worksheet, "sheet_state", "visible") or "visible"),
        "max_row": max_row,
        "max_column": max_column,
        "header_row": header_row,
        "headers": headers,
        "data_rows_scanned": data_rows_scanned,
        "non_empty_rows_scanned": non_empty_rows_scanned,
        "non_empty_cell_count_scanned": non_empty_cell_count,
        "formula_cell_count_scanned": formula_cell_count,
        "merged_cell_range_count": len(getattr(worksheet.merged_cells, "ranges", []) or []),
        "merged_cell_ranges_sample": merged_ranges,
        "table_count": len(tables),
        "tables": tables[:20],
        "chart_count": len(getattr(worksheet, "_charts", []) or []),
        "row_scan_capped": row_scan_capped,
        "column_scan_capped": column_scan_capped,
        "interpretation": (
            "XLSX workbook/sheet metadata for material preparation; not a statistical validation "
            "result and not verified figure provenance."
        ),
    }


def defined_name_count(workbook: Any) -> int:
    defined_names = getattr(workbook, "defined_names", None)
    if defined_names is None:
        return 0
    try:
        return len(list(defined_names.keys()))
    except Exception:  # noqa: BLE001 - openpyxl version compatibility.
        try:
            return len(list(defined_names))
        except Exception:  # noqa: BLE001
            return 0


def scan(root: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # noqa: BLE001 - dependency issue should be explicit.
        return {
            "schema_version": "0.1.0",
            "extractor": "scripts.xlsx_structure_extract",
            "scope_note": "XLSX structure intake requires openpyxl.",
            "input": {"package": str(root), "xlsx_files": 0},
            "xlsx_files": [],
            "sheets": [],
            "errors": [{
                "stage": "xlsx_structure_extraction",
                "error": f"openpyxl unavailable: {exc.__class__.__name__}",
            }],
        }

    xlsx_files: list[dict[str, Any]] = []
    sheets: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    for path in collect_xlsx_files(root):
        rel = str(path.relative_to(root))
        try:
            workbook = load_workbook(path, read_only=False, data_only=False)
        except Exception as exc:  # noqa: BLE001
            error = {
                "path": rel,
                "stage": "xlsx_structure_extraction",
                "error": str(exc),
            }
            xlsx_files.append({
                "path": rel,
                "sheet_count": 0,
                "visible_sheet_count": 0,
                "hidden_sheet_count": 0,
                "sheets_indexed": 0,
                "errors": [error],
            })
            errors.append(error)
            continue
        try:
            file_sheets: list[dict[str, Any]] = []
            for worksheet in workbook.worksheets:
                sheet_payload = scan_sheet(rel, worksheet)
                file_sheets.append(sheet_payload)
                sheets.append(sheet_payload)
            hidden_count = sum(1 for sheet in file_sheets if sheet.get("sheet_state") != "visible")
            xlsx_files.append({
                "path": rel,
                "sheet_count": len(workbook.worksheets),
                "visible_sheet_count": len(workbook.worksheets) - hidden_count,
                "hidden_sheet_count": hidden_count,
                "sheets_indexed": len(file_sheets),
                "defined_name_count": defined_name_count(workbook),
                "errors": [],
            })
        finally:
            workbook.close()

    return {
        "schema_version": "0.1.0",
        "extractor": "scripts.xlsx_structure_extract",
        "scope_note": (
            "Best-effort indexing of XLSX workbook and sheet metadata. This artifact records "
            "headers, sheet dimensions, formula/merged-cell counts, and figure/table-like labels "
            "for material preparation. It does not validate calculations, replace CSV exports, "
            "or prove source-to-figure provenance."
        ),
        "input": {
            "package": str(root),
            "xlsx_files": len(xlsx_files),
            "max_rows_scanned_per_sheet": MAX_ROWS_SCANNED,
            "max_columns_scanned_per_sheet": MAX_COLUMNS_SCANNED,
        },
        "xlsx_files": xlsx_files,
        "sheets": sheets,
        "errors": errors,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("package_dir", type=Path)
    parser.add_argument("--output", type=Path, default=Path("xlsx_structure.json"))
    args = parser.parse_args()

    root = args.package_dir.expanduser().resolve()
    if not root.exists() or not root.is_dir():
        raise SystemExit(f"Package directory not found: {root}")
    payload = scan(root)
    args.output.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({
        "output": str(args.output),
        "xlsx_files": payload["input"]["xlsx_files"],
        "sheets": len(payload["sheets"]),
        "errors": len(payload["errors"]),
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
