#!/usr/bin/env python3
"""Check biomedical source-data summaries for statistical consistency and weak forensic triage signals."""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any


CSV_EXTS = {".csv", ".tsv"}
XLSX_EXTS = {".xlsx"}
TABLE_EXTS = CSV_EXTS | XLSX_EXTS
NUMERIC_HINTS = ("mean", "sd", "sem", "se", "n", "p", "p_value", "pvalue")
TIME_HINT_RE = re.compile(r"(^|[_\-\s])(t\d+|day\d+|d\d+|week\d+|w\d+|hour\d+|h\d+|baseline|endpoint)($|[_\-\s])", re.I)
TIME_TOKEN_RE = re.compile(r"(t\d+|day\d+|d\d+|week\d+|w\d+|hour\d+|h\d+|baseline|endpoint)", re.I)
SUMMARY_COLUMNS = {"mean", "sd", "sem", "se", "n", "p", "p_value", "pvalue"}
TERMINAL_DIGIT_SKIP_COLUMNS = {"n"}
MISSING_NUMERIC_TEXT = {"na", "n/a", "nan", "null", "-"}
COUNT_HINTS = (
    "count",
    "counts",
    "integer",
    "event",
    "events",
    "cell",
    "cells",
    "colony",
    "colonies",
    "foci",
    "lesion",
    "lesions",
    "number",
)


def normalized_numeric_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if not text or text.lower() in MISSING_NUMERIC_TEXT:
        return None
    if text.startswith(("<", ">")):
        return None
    return text


def parse_float(value: Any) -> float | None:
    text = normalized_numeric_text(value)
    if text is None:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_header(header: str) -> str:
    return header.strip().lower().replace(" ", "_").replace("-", "_")


def cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def read_delimited_table(path: Path) -> list[dict[str, str]]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    with path.open(newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh, delimiter=delimiter)
        rows = []
        for row in reader:
            rows.append({normalize_header(k): v for k, v in row.items() if k is not None})
        return rows


def read_xlsx_tables(path: Path) -> list[tuple[Path, list[dict[str, str]]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # noqa: BLE001 - xlsx support is an explicit dependency.
        raise RuntimeError("XLSX source-data screening requires openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    tables: list[tuple[Path, list[dict[str, str]]]] = []
    for sheet in workbook.worksheets:
        headers: list[str] | None = None
        rows: list[dict[str, str]] = []
        for values in sheet.iter_rows(values_only=True):
            values = list(values)
            if headers is None:
                if not any(cell_to_text(value).strip() for value in values):
                    continue
                headers = [
                    normalize_header(cell_to_text(value)) if cell_to_text(value).strip() else f"column_{idx + 1}"
                    for idx, value in enumerate(values)
                ]
                continue
            if not any(cell_to_text(value).strip() for value in values):
                continue
            rows.append({
                header: cell_to_text(value)
                for header, value in zip(headers, values)
                if header
            })
        if rows:
            tables.append((Path(f"{path.name}#{sheet.title}"), rows))
    workbook.close()
    return tables


def read_tables(path: Path) -> list[tuple[Path, list[dict[str, str]]]]:
    if path.suffix.lower() in XLSX_EXTS:
        return read_xlsx_tables(path)
    return [(path, read_delimited_table(path))]


def row_label(path: Path, idx: int, row: dict[str, str]) -> str:
    for key in ("id", "animal_id", "mouse_id", "subject_id", "group", "condition", "figure", "panel", "comparison"):
        if row.get(key):
            return f"{path.name}:row{idx}:{key}={row[key]}"
    return f"{path.name}:row{idx}"


def risk_suggestion(level: str, evidence_type: str) -> str:
    if evidence_type == "weak_forensic_triage_signal" or level in {"R1", "R2"}:
        return "R2_max" if evidence_type == "weak_forensic_triage_signal" else f"{level}_possible"
    if level == "R4":
        return "R4_only_if_direct_contradiction"
    return f"{level}_possible"


def issue(location: str, risk_level: str, message: str, values: dict[str, Any], evidence_type: str = "statistics_consistency") -> dict[str, Any]:
    is_weak = evidence_type == "weak_forensic_triage_signal"
    return {
        "candidate_id": "",
        "detector": "stats.consistency_check",
        "candidate_type": "weak_statistical_signal" if is_weak else "statistical_consistency_candidate",
        "locations": [location],
        "finding_type": message,
        "evidence": {
            "message": message,
            "evidence_type": evidence_type,
            **values,
        },
        "evidence_strength": "weak_signal" if is_weak else "candidate",
        "risk_suggestion": risk_suggestion(risk_level, evidence_type),
        "risk_cap_tags": [evidence_type, "weak_statistical_signal", "weak_signal"] if is_weak else [evidence_type],
        "benign_explanations": [
            "rounding, normalization, export, or reporting differences may explain the observation",
            "source/raw records and analysis code are needed before escalation",
        ],
        "required_materials": [
            "source data",
            "raw records where applicable",
            "analysis file or code",
        ],
        "recommended_action": "Inspect source data, analysis code, rounding, normalization, and benign explanations.",
        "requires_contextual_calibration": True,
    }


def weak_issue(location: str, message: str, values: dict[str, Any]) -> dict[str, Any]:
    return issue(location, "R2", message, values, evidence_type="weak_forensic_triage_signal")


def almost_equal(a: float, b: float, tolerance: float) -> bool:
    return abs(a - b) <= max(tolerance, max(abs(a), abs(b)) * tolerance)


def row_suggests_integer_count(row: dict[str, str]) -> bool:
    text = " ".join([*row.keys(), *[str(v) for v in row.values()]]).lower()
    return any(hint in text for hint in COUNT_HINTS)


def near_integer(value: float, tolerance: float) -> bool:
    return abs(value - round(value)) <= tolerance


def integer_count_summary_issue(mean: float, sd: float | None, n: float, tolerance: float = 1e-3) -> dict[str, Any] | None:
    if n <= 0 or not near_integer(n, tolerance):
        return None
    n_int = int(round(n))
    total = mean * n_int
    if not near_integer(total, tolerance):
        return {
            "reason": "mean * n is not an integer for an integer-count outcome",
            "mean": mean,
            "n": n_int,
            "mean_times_n": total,
        }
    if sd is None or sd < 0 or n_int <= 1:
        return None

    total_int = round(total)
    sumsq_from_sample_sd = (sd ** 2) * (n_int - 1) + (total_int ** 2) / n_int
    sumsq_from_population_sd = (sd ** 2) * n_int + (total_int ** 2) / n_int
    relaxed = max(0.02, tolerance * max(1.0, abs(sumsq_from_sample_sd), abs(sumsq_from_population_sd)))
    sample_possible = near_integer(sumsq_from_sample_sd, relaxed)
    population_possible = near_integer(sumsq_from_population_sd, relaxed)
    if not sample_possible and not population_possible:
        return {
            "reason": "mean/SD/n do not imply an integer-valued sum of squares under sample or population SD",
            "mean": mean,
            "sd": sd,
            "n": n_int,
            "sum": total_int,
            "sample_sd_implied_sumsq": sumsq_from_sample_sd,
            "population_sd_implied_sumsq": sumsq_from_population_sd,
            "tolerance": relaxed,
        }
    return None


def check_rows(path: Path, rows: list[dict[str, str]], sem_tolerance: float) -> list[dict[str, Any]]:
    findings = []
    for idx, row in enumerate(rows, start=2):
        mean = parse_float(row.get("mean"))
        sd = parse_float(row.get("sd"))
        sem = parse_float(row.get("sem") or row.get("se"))
        n = parse_float(row.get("n"))
        p_value = parse_float(row.get("p") or row.get("p_value") or row.get("pvalue"))
        label = row_label(path, idx, row)

        if n is not None and (n <= 0 or abs(n - round(n)) > 1e-9):
            findings.append(issue(label, "R2", "n is non-positive or non-integer", {"n": n}))

        if sd is not None and sd < 0:
            findings.append(issue(label, "R3", "SD is negative", {"sd": sd}))
        if sem is not None and sem < 0:
            findings.append(issue(label, "R3", "SEM is negative", {"sem": sem}))

        if sd is not None and sem is not None and n is not None and n > 1:
            expected_sd = sem * math.sqrt(n)
            tolerance = max(sem_tolerance, abs(expected_sd) * sem_tolerance)
            if abs(sd - expected_sd) > tolerance:
                findings.append(issue(label, "R2", "SD is not consistent with SEM * sqrt(n)", {
                    "sd": sd,
                    "sem": sem,
                    "n": n,
                    "expected_sd_from_sem": expected_sd,
                }))
            if abs(sd - sem) <= tolerance and n > 2:
                findings.append(issue(label, "R2", "SD and SEM are nearly identical despite n > 2", {
                    "sd": sd,
                    "sem": sem,
                    "n": n,
                }))

        if p_value is not None and not (0 <= p_value <= 1):
            findings.append(issue(label, "R3", "p value is outside [0, 1]", {"p_value": p_value}))

        if mean is not None and sd is not None and abs(mean) > 0 and sd / abs(mean) < 1e-6:
            findings.append(issue(label, "R1", "Extremely small relative SD; weak triage signal", {
                "mean": mean,
                "sd": sd,
            }, evidence_type="weak_forensic_triage_signal"))

        if mean is not None and n is not None and row_suggests_integer_count(row):
            impossible = integer_count_summary_issue(mean, sd, n)
            if impossible:
                findings.append(issue(label, "R2", "Integer-count mean/SD/n combination appears mathematically incompatible", impossible))
    return findings


def numeric_columns(rows: list[dict[str, str]]) -> dict[str, list[tuple[int, str, float]]]:
    columns: dict[str, list[tuple[int, str, float]]] = defaultdict(list)
    for idx, row in enumerate(rows, start=2):
        for key, raw in row.items():
            value = parse_float(raw)
            if value is not None:
                columns[key].append((idx, str(raw), value))
    return dict(columns)


def terminal_digit(raw: str) -> str | None:
    text = normalized_numeric_text(raw)
    if text is None:
        return None
    text = text.lower()
    if "e" in text:
        return None
    digits = re.sub(r"[^0-9]", "", text)
    return digits[-1] if digits else None


def decimal_places(raw: str) -> int | None:
    text = normalized_numeric_text(raw)
    if text is None:
        return None
    text = text.lower()
    if "e" in text:
        return None
    if "." not in text:
        return 0
    return len(text.split(".", 1)[1].rstrip("%"))


def ones_digit(raw: str) -> str | None:
    text = normalized_numeric_text(raw)
    if text is None:
        return None
    text = text.lower()
    if "e" in text:
        return None
    before_decimal = text.split(".", 1)[0]
    digits = re.sub(r"[^0-9]", "", before_decimal)
    return digits[-1] if digits else None


def first_decimal_digit(raw: str) -> str | None:
    text = normalized_numeric_text(raw)
    if text is None:
        return None
    text = text.lower()
    if "e" in text or "." not in text:
        return None
    after_decimal = text.split(".", 1)[1]
    digits = re.sub(r"[^0-9]", "", after_decimal)
    return digits[0] if digits else None


def effective_min_count(n_values: int, user_min: int | None = None) -> int:
    if user_min is not None:
        return user_min
    if n_values <= 5:
        return 3
    if n_values <= 12:
        return 4
    return 8


def check_terminal_digits(path: Path, columns: dict[str, list[tuple[int, str, float]]], min_count: int | None, dominance: float) -> list[dict[str, Any]]:
    findings = []
    for column, values in columns.items():
        if column in TERMINAL_DIGIT_SKIP_COLUMNS:
            continue
        digits = [digit for _, raw, _ in values if (digit := terminal_digit(raw)) is not None]
        threshold = effective_min_count(len(digits), min_count)
        if len(digits) < threshold:
            continue
        counts = Counter(digits)
        digit, count = counts.most_common(1)[0]
        share = count / len(digits)
        if share >= dominance:
            findings.append(issue(f"{path.name}:{column}", "R1", "Terminal-digit preference; weak triage signal", {
                "column": column,
                "values_screened": len(digits),
                "effective_min_count": threshold,
                "dominant_terminal_digit": digit,
                "dominant_share": round(share, 3),
                "digit_counts": dict(sorted(counts.items())),
            }, evidence_type="weak_forensic_triage_signal"))
    return findings


def check_rounding_patterns(path: Path, columns: dict[str, list[tuple[int, str, float]]], min_count: int | None, share_threshold: float) -> list[dict[str, Any]]:
    findings = []
    for column, values in columns.items():
        if column in TERMINAL_DIGIT_SKIP_COLUMNS:
            continue
        raw_values = [raw for _, raw, _ in values]
        threshold = effective_min_count(len(raw_values), min_count)
        if len(raw_values) < threshold:
            continue
        terminal = [terminal_digit(raw) for raw in raw_values]
        terminal = [d for d in terminal if d is not None]
        threshold = effective_min_count(len(terminal), min_count)
        if len(terminal) < threshold:
            continue
        zero_or_five = sum(1 for digit in terminal if digit in {"0", "5"})
        decimal_counts = Counter(dp for raw in raw_values if (dp := decimal_places(raw)) is not None)
        if zero_or_five / len(terminal) >= share_threshold:
            findings.append(issue(f"{path.name}:{column}", "R1", "Values disproportionately end in 0 or 5; possible rounding/convenience pattern", {
                "column": column,
                "values_screened": len(terminal),
                "effective_min_count": threshold,
                "zero_or_five_share": round(zero_or_five / len(terminal), 3),
                "decimal_place_counts": dict(sorted(decimal_counts.items())),
            }, evidence_type="weak_forensic_triage_signal"))
    return findings


def check_precision_mixing(path: Path, columns: dict[str, list[tuple[int, str, float]]], min_count: int | None) -> list[dict[str, Any]]:
    findings = []
    dominant_places = {}
    column_counts = {}
    for column, values in columns.items():
        if column in TERMINAL_DIGIT_SKIP_COLUMNS:
            continue
        places = [dp for _, raw, _ in values if (dp := decimal_places(raw)) is not None]
        threshold = effective_min_count(len(places), min_count)
        if len(places) < threshold:
            continue
        counts = Counter(places)
        dominant, dominant_count = counts.most_common(1)[0]
        dominant_places[column] = dominant
        column_counts[column] = dict(sorted(counts.items()))
        if len(counts) > 1 and dominant_count / len(places) <= 0.8:
            findings.append(issue(f"{path.name}:{column}", "R1", "Mixed numeric precision within one column; weak triage signal", {
                "column": column,
                "values_screened": len(places),
                "effective_min_count": threshold,
                "decimal_place_counts": dict(sorted(counts.items())),
            }, evidence_type="weak_forensic_triage_signal"))

    if len(set(dominant_places.values())) >= 2:
        findings.append(issue(f"{path.name}:numeric_precision", "R1", "Numeric precision differs systematically across columns; weak triage signal", {
            "dominant_decimal_places_by_column": dominant_places,
            "decimal_place_counts_by_column": column_counts,
        }, evidence_type="weak_forensic_triage_signal"))
    return findings


def check_repeated_summaries(path: Path, rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    findings = []
    if not rows:
        return findings
    pair_counts: Counter[tuple[str, str]] = Counter()
    pair_locations: dict[tuple[str, str], list[str]] = defaultdict(list)
    for idx, row in enumerate(rows, start=2):
        mean_raw = row.get("mean")
        sd_raw = row.get("sd")
        if parse_float(mean_raw) is None or parse_float(sd_raw) is None:
            continue
        key = (str(mean_raw).strip(), str(sd_raw).strip())
        pair_counts[key] += 1
        pair_locations[key].append(row_label(path, idx, row))
    repeated = {
        f"mean={mean}, sd={sd}": locations
        for (mean, sd), count in pair_counts.items()
        if count >= 2
        for locations in [pair_locations[(mean, sd)]]
    }
    if repeated:
        findings.append(weak_issue(f"{path.name}:mean/sd", "Repeated mean/SD pairs across rows; weak-to-moderate triage signal", {
            "repeated_pairs": repeated,
            "rows_screened": len(rows),
        }))
    return findings


def shared_pairs(a_values: list[tuple[int, str, float]], b_values: list[tuple[int, str, float]]) -> list[tuple[int, float, float]]:
    a_by_row = {idx: value for idx, _, value in a_values}
    b_by_row = {idx: value for idx, _, value in b_values}
    return [(idx, a_by_row[idx], b_by_row[idx]) for idx in sorted(a_by_row.keys() & b_by_row.keys())]


def shared_raw_pairs(a_values: list[tuple[int, str, float]], b_values: list[tuple[int, str, float]]) -> list[tuple[int, str, str, float, float]]:
    a_by_row = {idx: (raw, value) for idx, raw, value in a_values}
    b_by_row = {idx: (raw, value) for idx, raw, value in b_values}
    pairs = []
    for idx in sorted(a_by_row.keys() & b_by_row.keys()):
        a_raw, a_value = a_by_row[idx]
        b_raw, b_value = b_by_row[idx]
        pairs.append((idx, a_raw, b_raw, a_value, b_value))
    return pairs


def ols_fit(pairs: list[tuple[int, float, float]]) -> dict[str, float] | None:
    if len(pairs) < 3:
        return None
    xs = [x for _, x, _ in pairs]
    ys = [y for _, _, y in pairs]
    mean_x = sum(xs) / len(xs)
    mean_y = sum(ys) / len(ys)
    ss_x = sum((x - mean_x) ** 2 for x in xs)
    ss_y = sum((y - mean_y) ** 2 for y in ys)
    if ss_x <= 0 or ss_y <= 0:
        return None
    cov = sum((x - mean_x) * (y - mean_y) for x, y in zip(xs, ys))
    slope = cov / ss_x
    intercept = mean_y - slope * mean_x
    residuals = [y - (slope * x + intercept) for x, y in zip(xs, ys)]
    sse = sum(r * r for r in residuals)
    rmse = math.sqrt(sse / len(residuals))
    max_abs = max(abs(r) for r in residuals)
    r2 = 1 - (sse / ss_y) if ss_y > 0 else 1.0
    scale = max(1.0, max(abs(x) for x in xs), max(abs(y) for y in ys))
    return {
        "slope": slope,
        "intercept": intercept,
        "rmse": rmse,
        "max_abs_residual": max_abs,
        "r2": r2,
        "scale": scale,
    }


def pearson(xs: list[float], ys: list[float]) -> float | None:
    if len(xs) != len(ys) or len(xs) < 3:
        return None
    mean_x = sum(xs) / len(xs)
    mean_y = sum(ys) / len(ys)
    ss_x = sum((x - mean_x) ** 2 for x in xs)
    ss_y = sum((y - mean_y) ** 2 for y in ys)
    if ss_x <= 0 or ss_y <= 0:
        return None
    cov = sum((x - mean_x) * (y - mean_y) for x, y in zip(xs, ys))
    return cov / math.sqrt(ss_x * ss_y)


def ranks(values: list[float]) -> list[int]:
    ordered = sorted(range(len(values)), key=lambda idx: (values[idx], idx))
    result = [0] * len(values)
    for rank, idx in enumerate(ordered):
        result[idx] = rank
    return result


def looks_time_column(header: str) -> bool:
    return bool(TIME_HINT_RE.search(header))


def time_token(header: str) -> str | None:
    match = TIME_TOKEN_RE.search(header)
    return match.group(1).lower() if match else None


def strip_time_token(header: str) -> str:
    stripped = TIME_TOKEN_RE.sub("", header)
    stripped = re.sub(r"[_\-\s]+", "_", stripped).strip("_")
    return stripped or header


def time_base(header: str) -> str:
    stripped = TIME_TOKEN_RE.sub("", header)
    stripped = re.sub(r"[_\-\s]+", "_", stripped).strip("_")
    return stripped or "__time_series__"


def relation_label(fit: dict[str, float], tolerance: float) -> str:
    slope = fit["slope"]
    intercept = fit["intercept"]
    scale = fit["scale"]
    if abs(slope - 1) <= 0.01 and abs(intercept) <= tolerance * scale:
        return "exact or near-exact duplicate column"
    if abs(slope - 1) <= 0.01 and abs(intercept) > tolerance * scale:
        return "whole-column additive/subtractive shift"
    if abs(intercept) <= tolerance * scale and abs(slope - 1) > 0.01:
        return "whole-column multiplicative/divisive scaling"
    return "whole-column affine linear transformation"


def check_column_relationships(
    path: Path,
    columns: dict[str, list[tuple[int, str, float]]],
    min_pairs: int,
    residual_tolerance: float,
) -> list[dict[str, Any]]:
    findings = []
    names = list(columns)
    for i, left in enumerate(names):
        for j, right in enumerate(names[i + 1:], start=i + 1):
            if left in SUMMARY_COLUMNS or right in SUMMARY_COLUMNS:
                continue
            left_time = time_token(left)
            right_time = time_token(right)
            if left_time and right_time:
                if left_time != right_time:
                    continue
            pairs = shared_pairs(columns[left], columns[right])
            if len(pairs) < min_pairs:
                continue
            xs = [x for _, x, _ in pairs]
            ys = [y for _, _, y in pairs]
            fit = ols_fit(pairs)
            if not fit:
                continue
            tolerance = residual_tolerance * fit["scale"]
            same_rank = ranks(xs) == ranks(ys)
            corr = pearson(xs, ys)
            near_exact_linear = fit["r2"] >= 0.999999 and fit["max_abs_residual"] <= tolerance

            if near_exact_linear:
                label = relation_label(fit, residual_tolerance)
                if looks_time_column(left) and looks_time_column(right):
                    message = f"Adjacent/paired time columns show {label}; weak-to-moderate triage signal"
                else:
                    message = f"Columns show {label}; weak-to-moderate triage signal"
                findings.append(weak_issue(f"{path.name}:{left}<->{right}", message, {
                    "left_column": left,
                    "right_column": right,
                    "paired_rows": len(pairs),
                    "slope": round(fit["slope"], 8),
                    "intercept": round(fit["intercept"], 8),
                    "r2": round(fit["r2"], 10),
                    "max_abs_residual": fit["max_abs_residual"],
                    "same_rank_order": same_rank,
                    "centered_residual_correlation": round(corr, 10) if corr is not None else None,
                }))
            elif same_rank and len(set(xs)) == len(xs) and len(set(ys)) == len(ys):
                findings.append(issue(f"{path.name}:{left}<->{right}", "R1", "Rank order is identical across columns; weak triage signal", {
                    "left_column": left,
                    "right_column": right,
                    "paired_rows": len(pairs),
                    "r2": round(fit["r2"], 6),
                }, evidence_type="weak_forensic_triage_signal"))
            elif corr is not None and abs(corr) >= 0.995:
                findings.append(weak_issue(f"{path.name}:{left}<->{right}", "Residual/rank fluctuation pattern is highly correlated across columns", {
                    "left_column": left,
                    "right_column": right,
                    "paired_rows": len(pairs),
                    "centered_residual_correlation": round(corr, 10),
                    "r2": round(fit["r2"], 6),
                }))
    return findings


def check_digit_preservation(
    path: Path,
    columns: dict[str, list[tuple[int, str, float]]],
    min_pairs: int,
    share_threshold: float,
) -> list[dict[str, Any]]:
    findings = []
    names = [name for name in columns if name not in SUMMARY_COLUMNS and name not in TERMINAL_DIGIT_SKIP_COLUMNS]
    digit_extractors = {
        "ones_digit": ones_digit,
        "first_decimal_digit": first_decimal_digit,
        "terminal_digit": terminal_digit,
    }
    for i, left in enumerate(names):
        for right in names[i + 1:]:
            raw_pairs = shared_raw_pairs(columns[left], columns[right])
            if len(raw_pairs) < min_pairs:
                continue
            preserved = {}
            for digit_name, extractor in digit_extractors.items():
                comparable = []
                for _, left_raw, right_raw, _, _ in raw_pairs:
                    left_digit = extractor(left_raw)
                    right_digit = extractor(right_raw)
                    if left_digit is not None and right_digit is not None:
                        comparable.append(left_digit == right_digit)
                if len(comparable) >= min_pairs:
                    share = sum(1 for matched in comparable if matched) / len(comparable)
                    if share >= share_threshold:
                        preserved[digit_name] = {
                            "comparable_pairs": len(comparable),
                            "match_share": round(share, 3),
                        }
            if preserved:
                findings.append(weak_issue(f"{path.name}:{left}<->{right}", "Digit positions are preserved across paired columns", {
                    "left_column": left,
                    "right_column": right,
                    "paired_rows": len(raw_pairs),
                    "preserved_digit_positions": preserved,
                }))
    return findings


def check_time_stratified_shifts(
    path: Path,
    columns: dict[str, list[tuple[int, str, float]]],
    min_pairs: int,
    residual_tolerance: float,
) -> list[dict[str, Any]]:
    findings = []
    grouped_offsets: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    names = [name for name in columns if name not in SUMMARY_COLUMNS]
    for i, left in enumerate(names):
        left_time = time_token(left)
        if left_time is None:
            continue
        for right in names[i + 1:]:
            right_time = time_token(right)
            if left_time != right_time:
                continue
            pairs = shared_pairs(columns[left], columns[right])
            if len(pairs) < min_pairs:
                continue
            fit = ols_fit(pairs)
            if not fit:
                continue
            tolerance = residual_tolerance * fit["scale"]
            if abs(fit["slope"] - 1) <= 0.01 and abs(fit["intercept"]) > tolerance and fit["max_abs_residual"] <= tolerance:
                left_base = time_base(left)
                right_base = time_base(right)
                key = tuple(sorted((left_base, right_base)))
                grouped_offsets[key].append({
                    "time": left_time,
                    "left_column": left,
                    "right_column": right,
                    "offset": round(fit["intercept"], 8),
                    "paired_rows": len(pairs),
                })

    for (left_base, right_base), offsets in grouped_offsets.items():
        if len(offsets) < 2:
            continue
        unique_offsets = sorted({item["offset"] for item in offsets})
        findings.append(weak_issue(f"{path.name}:{left_base}<->{right_base}", "Time-stratified additive shifts across group columns", {
            "group_column_bases": [left_base, right_base],
            "timepoint_offsets": offsets,
            "unique_offsets": unique_offsets,
        }))
    return findings


def row_identity(row: dict[str, str], idx: int) -> str:
    for key in ("animal_id", "mouse_id", "subject_id", "sample_id", "id"):
        if row.get(key):
            return str(row[key])
    return f"row{idx}"


def time_ordered_numeric_column_groups(rows: list[dict[str, str]], columns: dict[str, list[tuple[int, str, float]]]) -> list[list[str]]:
    if not rows:
        return []
    names = [name for name in rows[0] if name in columns and len(columns[name]) >= 3]
    time_like = [name for name in names if looks_time_column(name)]
    if len(time_like) >= 3:
        groups: dict[str, list[str]] = defaultdict(list)
        for name in time_like:
            groups[time_base(name)].append(name)
        return [group for group in groups.values() if len(group) >= 3]
    return [names] if len(names) >= 4 else []


def check_longitudinal_mechanics(path: Path, rows: list[dict[str, str]], columns: dict[str, list[tuple[int, str, float]]], tolerance: float) -> list[dict[str, Any]]:
    findings = []
    for time_cols in time_ordered_numeric_column_groups(rows, columns):
        if len(time_cols) < 4:
            continue

        linear_rows = []
        increment_patterns: dict[tuple[float, ...], list[str]] = defaultdict(list)
        for idx, row in enumerate(rows, start=2):
            values = [parse_float(row.get(col)) for col in time_cols]
            if any(value is None for value in values):
                continue
            series = [float(value) for value in values if value is not None]
            diffs = [series[i + 1] - series[i] for i in range(len(series) - 1)]
            if not diffs:
                continue
            scale = max(1.0, max(abs(value) for value in series))
            mean_diff = sum(diffs) / len(diffs)
            max_deviation = max(abs(diff - mean_diff) for diff in diffs)
            label = row_identity(row, idx)
            pattern = tuple(round(diff, 4) for diff in diffs)
            increment_patterns[pattern].append(label)
            if max_deviation <= tolerance * scale:
                linear_rows.append({
                    "row": label,
                    "increments": [round(diff, 6) for diff in diffs],
                    "max_increment_deviation": max_deviation,
                })

        screened = len([row for row in rows if all(parse_float(row.get(col)) is not None for col in time_cols)])
        if screened >= 3 and len(linear_rows) / screened >= 0.6:
            findings.append(weak_issue(f"{path.name}:{','.join(time_cols)}", "Longitudinal trajectories are unusually linear/mechanical across rows", {
                "time_columns": time_cols,
                "rows_screened": screened,
                "linear_rows": linear_rows[:12],
                "linear_row_share": round(len(linear_rows) / screened, 3),
            }))

        repeated_patterns = {str(pattern): labels for pattern, labels in increment_patterns.items() if len(labels) >= 3}
        if repeated_patterns:
            findings.append(weak_issue(f"{path.name}:{','.join(time_cols)}", "Repeated longitudinal increment pattern across animals/samples", {
                "time_columns": time_cols,
                "repeated_increment_patterns": repeated_patterns,
            }))
    return findings


def check_table_forensics(
    path: Path,
    rows: list[dict[str, str]],
    min_pairs: int,
    min_digit_count: int | None,
    digit_dominance: float,
    rounding_share: float,
    residual_tolerance: float,
) -> list[dict[str, Any]]:
    columns = numeric_columns(rows)
    findings = []
    findings.extend(check_terminal_digits(path, columns, min_digit_count, digit_dominance))
    findings.extend(check_rounding_patterns(path, columns, min_digit_count, rounding_share))
    findings.extend(check_precision_mixing(path, columns, min_digit_count))
    findings.extend(check_repeated_summaries(path, rows))
    findings.extend(check_column_relationships(path, columns, min_pairs, residual_tolerance))
    findings.extend(check_digit_preservation(path, columns, min_pairs, 0.85))
    findings.extend(check_time_stratified_shifts(path, columns, min_pairs, residual_tolerance))
    findings.extend(check_longitudinal_mechanics(path, rows, columns, residual_tolerance))
    return findings


def column_sequence(values: list[tuple[int, str, float]]) -> tuple[float, ...] | None:
    if not values:
        return None
    return tuple(round(value, 10) for _, _, value in values)


def check_cross_file_sequence_reuse(tables: list[tuple[Path, list[dict[str, str]], dict[str, list[tuple[int, str, float]]]]], min_pairs: int) -> list[dict[str, Any]]:
    sequence_locations: dict[tuple[float, ...], list[dict[str, Any]]] = defaultdict(list)
    for path, _rows, columns in tables:
        for column, values in columns.items():
            if column in SUMMARY_COLUMNS or len(values) < min_pairs:
                continue
            sequence = column_sequence(values)
            if sequence is None:
                continue
            sequence_locations[sequence].append({
                "file": path.name,
                "column": column,
                "rows": [idx for idx, _, _ in values],
            })

    findings = []
    for sequence, locations in sequence_locations.items():
        files = {item["file"] for item in locations}
        if len(locations) >= 2 and len(files) >= 2:
            findings.append(weak_issue("cross-file numeric sequences", "Identical numeric sequence appears in multiple files/figures", {
                "sequence_length": len(sequence),
                "sequence_preview": list(sequence[:8]),
                "locations": locations,
            }))
    return findings


def collect_files(path: Path) -> list[Path]:
    if path.is_file():
        return [path] if path.suffix.lower() in TABLE_EXTS else []
    return [p for p in sorted(path.rglob("*")) if p.is_file() and p.suffix.lower() in TABLE_EXTS]


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("path", type=Path, help="CSV/TSV file or folder containing source-data tables")
    parser.add_argument("--sem-tolerance", type=float, default=1e-3)
    parser.add_argument("--min-pairs", type=int, default=4, help="Minimum paired numeric rows for column relationship checks")
    parser.add_argument("--min-digit-count", type=int, default=None, help="Override adaptive minimum numeric values for terminal-digit and rounding checks")
    parser.add_argument("--digit-dominance", type=float, default=0.65, help="Dominant terminal digit share needed to flag")
    parser.add_argument("--rounding-share", type=float, default=0.85, help="Share of values ending in 0 or 5 needed to flag")
    parser.add_argument("--residual-tolerance", type=float, default=1e-9, help="Relative tolerance for exact linear transformation screens")
    parser.add_argument("--output", type=Path, default=Path("stats_consistency_findings.json"))
    args = parser.parse_args()

    target = args.path.expanduser().resolve()
    files = collect_files(target)
    candidates = []
    errors = []
    tables = []
    for file_path in files:
        try:
            for table_path, rows in read_tables(file_path):
                columns = numeric_columns(rows)
                tables.append((table_path, rows, columns))
                candidates.extend(check_rows(table_path, rows, args.sem_tolerance))
                candidates.extend(check_table_forensics(
                    table_path,
                    rows,
                    args.min_pairs,
                    args.min_digit_count,
                    args.digit_dominance,
                    args.rounding_share,
                    args.residual_tolerance,
                ))
        except Exception as exc:  # noqa: BLE001 - report unreadable data without aborting.
            errors.append({"path": str(file_path), "error": str(exc)})
    candidates.extend(check_cross_file_sequence_reuse(tables, args.min_pairs))
    for idx, item in enumerate(candidates, start=1):
        item["candidate_id"] = f"BIOMED-STAT-{idx:04d}"
    result = {
        "detector_name": "stats.consistency_check",
        "detector_version": "0.3.0",
        "input": {
            "path": str(target),
            "sem_tolerance": args.sem_tolerance,
            "min_pairs": args.min_pairs,
            "min_digit_count": args.min_digit_count,
            "digit_dominance": args.digit_dominance,
            "rounding_share": args.rounding_share,
            "residual_tolerance": args.residual_tolerance,
        },
        "files_screened": [str(p) for p in files],
        "candidates": candidates,
        "errors": errors,
    }
    args.output.write_text(json.dumps(result, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({
        "output": str(args.output),
        "files_screened": len(files),
        "candidates": len(candidates),
        "errors": len(errors),
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
